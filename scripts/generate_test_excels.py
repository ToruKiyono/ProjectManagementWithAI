from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parent.parent


REQUIREMENT_HEADERS = [
    "编号", "类型", "附加类别", "标题", "父工作项编号", "状态", "类别", "来源渠道", "归属空间编号", "归属",
    "迭代计划", "责任人", "创建人", "标签", "描述", "测试责任人", "参与人", "归属微服务", "发布计划", "优先级",
    "验证策略", "验证策略描述", "是否内容合规", "是否涉及安全", "创建时间", "开始时间", "完成时间", "关闭时间", "分发RR编号", "关联设计文档",
    "关联MR", "设计策略", "会签", "设计策略描述", "备注", "工作量 人·天", "工作量 人·月", "所属特性集", "进度", "进度百分比",
    "特性属性", "需求来源", "客户", "代码行数", "Offering", "发现问题数量", "昇腾计算依赖需求JDC", "SE", "风险", "检查点",
    "用例个数", "资料工作量（人天）", "是否周边依赖，如有，请描述", "UCD工作量（人天）", "需求分类", "规划上线时间", "需求场景", "研发进展", "前端工作量（人天）"
]

ISSUE_HEADERS = [
    "问题编号", "发现环境", "挂起/撤销", "问题标题", "问题描述", "提出人", "提出方", "发现服务", "期望修复时间", "责任服务",
    "发现问题版本", "交付场景", "归属组织", "修复计划", "发布计划", "迭代计划", "迭代", "责任人", "当前责任人", "问题状态",
    "问题阶段", "严重程度", "问题类别", "发现站点", "影响范围", "发现阶段", "创建人", "发现方式", "发现版本", "发现计划",
    "发现迭代", "协同计划", "协同迭代", "标签", "研发责任人", "测试责任人", "健康度", "重现概率", "优先级", "预计工作量(人月)",
    "预计工作量(人天)", "预计工作量(代码量千行)", "提交滞留时间", "定位中滞留时间", "修复滞留时间", "待修复滞留时间", "修改中滞留时间", "测试滞留时间", "修复完成滞留时间", "待验收滞留时间",
    "退回待确认滞留时间", "问题创建时间", "发现时间", "接纳时间", "最后更新时间", "承诺修复时间", "开始修复时间", "转测时间", "修复完成（测试通过）", "上线时间",
    "延期天数", "问题关闭时间", "问题关闭人", "驳回次数", "驳回详情", "关联工作项", "源问题单", "同步问题单", "关闭方式", "代码合入次数",
    "结论操作人", "结论操作时间", "结论简述", "结论描述", "问题根因", "普通评论", "关键节点评论", "转测操作人", "备注", "评审结论",
    "评审五要素", "CCB裁决结论时间", "重复问题单", "特性编码", "是否特性引入", "解决计划", "遗留问题评审：原因分析", "遗留问题评审：规避措施", "遗留问题评审：用户影响", "团队",
    "用例来源系统", "用例来源系统编号", "用例来源系统链接", "来源系统", "来源系统编号", "来源系统链接"
]

TIMELINE_HEADERS = [
    "版本线", "版本号", "周时间", "日期", "小迭代", "说明", "分支", "toolkit", "开始时间", "结束时间"
]


def req_defaults() -> Dict[str, object]:
    return {header: "" for header in REQUIREMENT_HEADERS}


def issue_defaults() -> Dict[str, object]:
    return {header: "" for header in ISSUE_HEADERS}


def timeline_defaults() -> Dict[str, object]:
    return {header: "" for header in TIMELINE_HEADERS}


def build_requirements() -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []

    def add(
        req_id: str,
        title: str,
        version_line: str,
        release_plan: str,
        iteration_plan: str,
        owner: str,
        tester: str,
        progress_percent: int,
        status: str,
        rd_status: str,
        risk: str,
        online_plan: str,
        issue_count: int,
        feature_set: str,
        service: str,
        category: str,
        tags: str,
        priority: str,
        frontend_days: float,
        extra: Dict[str, object] | None = None,
    ) -> None:
        row = req_defaults()
        row.update(
            {
                "编号": req_id,
                "类型": "需求",
                "附加类别": "版本特性",
                "标题": title,
                "父工作项编号": "",
                "状态": status,
                "类别": category,
                "来源渠道": "产品规划会",
                "归属空间编号": "SPACE-AI-PLATFORM",
                "归属": feature_set,
                "迭代计划": iteration_plan,
                "责任人": owner,
                "创建人": "陈晓明",
                "标签": tags,
                "描述": f"{title}，用于支撑{version_line}版本线本轮交付目标，数据、日志与发布节奏均已纳入联调范围。",
                "测试责任人": tester,
                "参与人": f"{owner}、{tester}、李海涛",
                "归属微服务": service,
                "发布计划": release_plan,
                "优先级": priority,
                "验证策略": "自动化回归+冒烟验证",
                "验证策略描述": "覆盖关键主链路、灰度发布校验与回归用例抽检。",
                "是否内容合规": "是",
                "是否涉及安全": "是" if "鉴权" in title or "下载" in title else "否",
                "创建时间": "2026-02-24",
                "开始时间": "2026-02-26",
                "完成时间": "" if progress_percent < 95 else online_plan,
                "关闭时间": online_plan if status == "已上线" else "",
                "分发RR编号": f"RR-{req_id[-4:]}",
                "关联设计文档": f"design/{req_id.lower()}.md",
                "关联MR": f"!{1000 + len(rows) + 1}",
                "设计策略": "增量设计",
                "会签": "已完成" if status not in {"待处理", "设计中"} else "进行中",
                "设计策略描述": "优先保证主流程打通，再补充场景化兼容与观测指标。",
                "备注": "用于版本管理驾驶舱导入测试",
                "工作量 人·天": round(frontend_days + 4.0, 1),
                "工作量 人·月": round((frontend_days + 4.0) / 22, 2),
                "所属特性集": feature_set,
                "进度": f"{progress_percent}%",
                "进度百分比": progress_percent,
                "特性属性": "增量增强",
                "需求来源": "版本例会",
                "客户": "华北区域云业务部",
                "代码行数": 1200 + len(rows) * 230,
                "Offering": version_line,
                "发现问题数量": issue_count,
                "昇腾计算依赖需求JDC": f"JDC-{2600 + len(rows)}" if "训练" in title or "模型" in title else "",
                "SE": "李海涛",
                "风险": risk,
                "检查点": "设计评审/开发自测/联调转测",
                "用例个数": 18 + len(rows) * 3,
                "资料工作量（人天）": round(0.5 + len(rows) * 0.1, 1),
                "是否周边依赖，如有，请描述": "依赖模型仓库、IAM 与部署控制台接口",
                "UCD工作量（人天）": round(0.3 + len(rows) * 0.1, 1),
                "需求分类": category,
                "规划上线时间": online_plan,
                "需求场景": version_line,
                "研发进展": rd_status,
                "前端工作量（人天）": frontend_days,
            }
        )
        if extra:
            row.update(extra)
        rows.append(row)

    add(
        "REQ-HC-753-B001-001",
        "【预置模型】新增Qwen3.5等模型支持训练",
        "HC",
        "7.5.3",
        "7.5.3-B001",
        "王俊",
        "刘静",
        35,
        "设计中",
        "设计中",
        "设计会签未完成",
        "2026-03-09",
        2,
        "训练平台",
        "training-service",
        "模型训练",
        "预置模型,训练,Qwen3.5",
        "P0",
        2.5,
    )
    add(
        "REQ-HC-753-B001-002",
        "训练任务日志可观测性增强",
        "HC",
        "7.5.3",
        "7.5.3-B001",
        "张蕾",
        "赵婷",
        15,
        "待处理",
        "待处理",
        "",
        "2026-03-09",
        1,
        "训练平台",
        "log-observer",
        "可观测性",
        "日志,可观测,训练",
        "P2",
        1.5,
        {"是否涉及安全": "否", "是否周边依赖，如有，请描述": "依赖日志聚合与对象存储检索接口"},
    )
    add(
        "REQ-HC-753-B002-001",
        "推理2.0新增部署模式",
        "HC",
        "7.5.3",
        "7.5.3-B002",
        "周凯",
        "朱敏",
        62,
        "开发中",
        "开发中",
        "进度滞后",
        "2026-03-13",
        3,
        "推理2.0",
        "inference-gateway",
        "推理部署",
        "推理2.0,部署,弹性扩缩",
        "P0",
        3.0,
    )
    add(
        "REQ-HC-753-B003-001",
        "模型仓库升级与缓存优化",
        "HC",
        "7.5.3",
        "7.5.3-B003",
        "何松",
        "赵婷",
        88,
        "转测中",
        "转测中",
        "",
        "2026-03-18",
        2,
        "Common",
        "model-repo",
        "仓库能力",
        "模型仓库,缓存,下载",
        "P1",
        1.0,
    )
    add(
        "REQ-HC-760-B001-001",
        "多租户版本驾驶舱筛选性能优化",
        "HC",
        "7.6.0",
        "7.6.0-B001",
        "刘畅",
        "冯璐",
        55,
        "开发中",
        "开发中",
        "无",
        "2026-03-30",
        1,
        "UI",
        "dashboard-ui",
        "管理驾驶舱",
        "驾驶舱,筛选,性能",
        "P1",
        4.0,
        {"创建时间": "2026-03-02", "开始时间": "2026-03-04"},
    )
    add(
        "REQ-HC-760-B002-001",
        "版本健康度看板支持按团队聚合",
        "HC",
        "7.6.0",
        "7.6.0-B002",
        "刘畅",
        "冯璐",
        100,
        "已上线",
        "已上线",
        "",
        "2026-04-08",
        0,
        "UI",
        "dashboard-ui",
        "管理驾驶舱",
        "驾驶舱,健康度,团队",
        "P2",
        2.0,
        {"创建时间": "2026-03-06", "开始时间": "2026-03-08"},
    )
    add(
        "REQ-HCS-861-B071-001",
        "【HCS】AZ容灾场景下在线服务编排增强",
        "HCS",
        "8.6.1",
        "8.6.1-B071",
        "罗威",
        "秦雪",
        78,
        "转测中",
        "转测中",
        "转测阻塞",
        "2026-03-12",
        4,
        "推理2.0",
        "service-orchestrator",
        "高可用",
        "HCS,AZ容灾,推理2.0",
        "P0",
        2.5,
        {"创建时间": "2026-02-27", "开始时间": "2026-03-01"},
    )
    add(
        "REQ-HCS-861-B071-002",
        "HCS租户隔离下服务列表查询链路加固",
        "HCS",
        "8.6.1",
        "8.6.1-B071",
        "高翔",
        "秦雪",
        42,
        "开发中",
        "开发中",
        "",
        "2026-03-12",
        2,
        "Common",
        "iam-proxy",
        "权限链路",
        "HCS,租户隔离,查询",
        "P1",
        1.0,
        {"是否涉及安全": "是"},
    )
    add(
        "REQ-HCS-861-B072-001",
        "推理服务灰度回滚策略补齐",
        "HCS",
        "8.6.1",
        "8.6.1-B072",
        "罗威",
        "秦雪",
        92,
        "问题单清理",
        "问题单清理",
        "无",
        "2026-03-19",
        1,
        "推理2.0",
        "service-orchestrator",
        "发布管控",
        "HCS,灰度,回滚",
        "P1",
        1.5,
    )
    add(
        "REQ-HCSO-750-B961-001",
        "HCSO环境下模型下载链路优化",
        "HCSO",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B961",
        "陈斌",
        "韩梅",
        68,
        "开发中",
        "开发中",
        "进度滞后",
        "2026-03-09",
        2,
        "Common",
        "artifact-proxy",
        "离线交付",
        "HCSO,下载,离线包",
        "P0",
        1.5,
        {"创建时间": "2026-02-20", "开始时间": "2026-02-25", "是否涉及安全": "是"},
    )
    add(
        "REQ-HCSO-750-B962-001",
        "离线包校验工具支持断点续传",
        "HCSO",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B962",
        "宋健",
        "韩梅",
        100,
        "已上线",
        "已上线",
        "",
        "2026-03-16",
        0,
        "工具",
        "offline-toolkit",
        "离线交付",
        "HCSO,工具,断点续传",
        "P2",
        0.5,
        {"创建时间": "2026-02-22", "开始时间": "2026-02-26"},
    )
    return rows


def build_issues() -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []

    def add(
        issue_id: str,
        title: str,
        description: str,
        scene: str,
        found_version: str,
        found_iteration: str,
        issue_status: str,
        issue_stage: str,
        severity: str,
        team: str,
        linked_req: str,
        owner: str,
        tester: str,
        release_plan: str,
        iteration_plan: str,
        repair_plan: str,
        found_time: str,
        due_time: str,
        extra: Dict[str, object] | None = None,
    ) -> None:
        row = issue_defaults()
        row.update(
            {
                "问题编号": issue_id,
                "发现环境": "SIT",
                "挂起/撤销": "",
                "问题标题": title,
                "问题描述": description,
                "提出人": tester,
                "提出方": "测试",
                "发现服务": team,
                "期望修复时间": due_time,
                "责任服务": team,
                "发现问题版本": found_iteration or found_version,
                "交付场景": scene,
                "归属组织": "云平台AI产品部",
                "修复计划": repair_plan,
                "发布计划": release_plan,
                "迭代计划": iteration_plan,
                "迭代": iteration_plan,
                "责任人": owner,
                "当前责任人": owner,
                "问题状态": issue_status,
                "问题阶段": issue_stage,
                "严重程度": severity,
                "问题类别": "功能",
                "发现站点": "西安测试中心",
                "影响范围": "主流程" if severity in {"严重", "致命"} else "局部功能",
                "发现阶段": "系统测试",
                "创建人": tester,
                "发现方式": "手工执行用例",
                "发现版本": found_version,
                "发现计划": release_plan,
                "发现迭代": found_iteration,
                "协同计划": release_plan,
                "协同迭代": iteration_plan,
                "标签": f"{scene},{team},联调",
                "研发责任人": owner,
                "测试责任人": tester,
                "健康度": "风险" if severity in {"严重", "致命"} else "关注",
                "重现概率": "高" if severity in {"严重", "致命"} else "中",
                "优先级": "P0" if severity == "致命" else ("P1" if severity == "严重" else "P2"),
                "预计工作量(人月)": 0.12 if severity in {"严重", "致命"} else 0.05,
                "预计工作量(人天)": 2.5 if severity in {"严重", "致命"} else 1.0,
                "预计工作量(代码量千行)": 0.8 if team in {"推理2.0", "训练", "Common"} else 0.2,
                "提交滞留时间": 0,
                "定位中滞留时间": 1 if issue_stage == "定位中" else 0,
                "修复滞留时间": 2 if issue_status in {"修复", "修复完成"} else 0,
                "待修复滞留时间": 0,
                "修改中滞留时间": 2 if issue_stage == "修改中" else 0,
                "测试滞留时间": 1 if issue_stage == "已转测" else 0,
                "修复完成滞留时间": 0,
                "待验收滞留时间": 0,
                "退回待确认滞留时间": 0,
                "问题创建时间": found_time,
                "发现时间": found_time,
                "接纳时间": found_time if issue_status in {"接纳", "修复", "修复完成", "关闭"} else "",
                "最后更新时间": due_time,
                "承诺修复时间": due_time,
                "开始修复时间": found_time if issue_status in {"修复", "修复完成", "关闭"} else "",
                "转测时间": due_time if issue_stage == "已转测" else "",
                "修复完成（测试通过）": due_time if issue_status in {"修复完成", "关闭"} else "",
                "上线时间": due_time if issue_status == "关闭" else "",
                "延期天数": 0,
                "问题关闭时间": due_time if issue_status == "关闭" else "",
                "问题关闭人": owner if issue_status == "关闭" else "",
                "驳回次数": 0,
                "驳回详情": "",
                "关联工作项": linked_req,
                "源问题单": "",
                "同步问题单": "",
                "关闭方式": "修复关闭" if issue_status == "关闭" else "",
                "代码合入次数": 1 if issue_status in {"修复", "修复完成", "关闭"} else 0,
                "结论操作人": owner if issue_status in {"修复完成", "关闭"} else "",
                "结论操作时间": due_time if issue_status in {"修复完成", "关闭"} else "",
                "结论简述": "已定位并完成修复验证" if issue_status in {"修复完成", "关闭"} else "问题受理中",
                "结论描述": "问题经过日志与接口排查后已明确根因，修复方案进入执行并跟踪回归结果。",
                "问题根因": "边界条件未覆盖，错误处理路径缺失",
                "普通评论": "导入测试示例数据",
                "关键节点评论": "保留不同状态与阶段组合供前端映射测试",
                "转测操作人": owner if issue_stage == "已转测" else "",
                "备注": "用于版本管理驾驶舱问题单导入",
                "评审结论": "通过" if issue_status in {"修复完成", "关闭"} else "待跟踪",
                "评审五要素": "版本/场景/责任人/计划/风险",
                "CCB裁决结论时间": due_time if severity == "致命" else "",
                "重复问题单": "",
                "特性编码": linked_req.replace("REQ", "FEAT") if linked_req else "",
                "是否特性引入": "是" if linked_req else "否",
                "解决计划": repair_plan,
                "遗留问题评审：原因分析": "",
                "遗留问题评审：规避措施": "",
                "遗留问题评审：用户影响": "",
                "团队": team,
                "用例来源系统": "TestLink",
                "用例来源系统编号": f"TL-{1200 + len(rows)}",
                "用例来源系统链接": f"https://testlink.local/case/{1200 + len(rows)}",
                "来源系统": "禅道",
                "来源系统编号": f"BUG-{3200 + len(rows)}",
                "来源系统链接": f"https://zentao.local/bug/{3200 + len(rows)}",
            }
        )
        if extra:
            row.update(extra)
        rows.append(row)

    add(
        "ISSUE-HC-753-B001-001",
        "【HC 7.5.3-B001】【训练】训练任务提交后日志页面空白",
        "测试在创建训练任务后进入日志页，页面仅展示空白容器且无错误提示。浏览器控制台可见日志查询接口返回 200，但前端未渲染返回内容。相同任务在旧版本日志页可以正常展示，说明问题集中在本轮页面渲染改造。",
        "HC",
        "7.5.3",
        "7.5.3-B001",
        "修复",
        "修改中",
        "严重",
        "训练",
        "REQ-HC-753-B001-002",
        "张蕾",
        "赵婷",
        "7.5.3",
        "7.5.3-B001",
        "7.5.3-B001",
        "2026-03-07",
        "2026-03-10",
    )
    add(
        "ISSUE-HC-753-B002-001",
        "【HC 7.5.3-B002】【推理2.0】新增部署模式下实例副本数未按策略生效",
        "在弹性部署模式中提交 4 副本策略，服务实际仅拉起 2 个实例。控制面显示策略创建成功，但实例编排日志中存在默认值回退。问题会导致部署能力验收无法继续，需要尽快明确是编排逻辑还是配置下发异常。",
        "HC",
        "7.5.3",
        "7.5.3-B002",
        "接纳",
        "定位中",
        "一般",
        "推理2.0",
        "REQ-HC-753-B002-001",
        "周凯",
        "朱敏",
        "7.5.3",
        "7.5.3-B002",
        "7.5.3-B002",
        "2026-03-11",
        "2026-03-14",
    )
    add(
        "ISSUE-HC-753-B003-001",
        "【HC 7.5.3-B003】【Common】模型仓库缓存命中后下载耗时仍异常偏高",
        "缓存命中场景下重复下载同一模型，链路耗时仍接近首次下载时间。压测日志显示缓存命中标记已生效，但文件流仍回源对象存储。问题已稳定复现，影响模型仓库优化收益评估。",
        "HC",
        "7.5.3",
        "7.5.3-B003",
        "修复完成",
        "已转测",
        "一般",
        "Common",
        "REQ-HC-753-B003-001",
        "何松",
        "赵婷",
        "7.5.3",
        "7.5.3-B003",
        "7.5.3-B003",
        "2026-03-17",
        "2026-03-20",
    )
    add(
        "ISSUE-HC-753-MAJOR-001",
        "【HC 7.5.3】【UI】版本概览页筛选条件切换后统计卡片未刷新",
        "测试在版本概览页连续切换团队与状态筛选项后，统计卡片数值保持上一次结果。接口返回数据正确，但页面缓存未刷新。该问题不依赖小迭代，可用于验证前端对仅大版本问题单的导入处理。",
        "HC",
        "7.5.3",
        "",
        "关闭",
        "关闭",
        "提示",
        "UI",
        "",
        "刘畅",
        "冯璐",
        "7.5.3",
        "7.5.3",
        "7.5.3",
        "2026-03-12",
        "2026-03-15",
        {"关闭方式": "配置修正关闭", "代码合入次数": 0},
    )
    add(
        "ISSUE-HCS-861-B071-001",
        "【HCS 8.6.1-B071】【推理2.0】AZ容灾场景下在线服务列表查询失败",
        "在双 AZ 容灾环境中，服务已成功部署但列表查询接口返回 500。接口日志提示跨 AZ 编排结果聚合为空，前端页面因此无法展示任何在线服务。该问题会阻塞 B071 场景化联调，是当前最高优先级问题之一。",
        "HCS",
        "8.6.1",
        "8.6.1-B071",
        "修复",
        "修改中",
        "致命",
        "推理2.0",
        "REQ-HCS-861-B071-001",
        "罗威",
        "秦雪",
        "8.6.1",
        "8.6.1-B071",
        "8.6.1-B071",
        "2026-03-12",
        "2026-03-13",
    )
    add(
        "ISSUE-HCS-861-B071-002",
        "【HCS 8.6.1-B071】【Common】租户隔离场景下鉴权透传头缺失",
        "使用子租户账号查询服务列表时，网关未透传租户鉴权头，导致下游统一返回无权限。抓包可见请求在入口层已经丢失关键 Header。问题涉及 IAM 代理和网关拼装逻辑，需要联合排查。",
        "HCS",
        "8.6.1",
        "8.6.1-B071",
        "接纳",
        "定位中",
        "严重",
        "Common",
        "REQ-HCS-861-B071-002",
        "高翔",
        "秦雪",
        "8.6.1",
        "8.6.1-B071",
        "8.6.1-B071",
        "2026-03-12",
        "2026-03-15",
    )
    add(
        "ISSUE-HCS-861-B072-001",
        "【HCS 8.6.1-B072】【推理2.0】灰度回滚后旧实例未自动摘流",
        "在灰度回滚流程中，服务实例已恢复旧版本镜像，但流量仍部分命中新实例。转测环境重复执行三次均可复现，说明摘流清理逻辑存在遗漏。问题修复后已重新提交测试验证。",
        "HCS",
        "8.6.1",
        "8.6.1-B072",
        "修复完成",
        "已转测",
        "一般",
        "推理2.0",
        "REQ-HCS-861-B072-001",
        "罗威",
        "秦雪",
        "8.6.1",
        "8.6.1-B072",
        "8.6.1-B072",
        "2026-03-18",
        "2026-03-21",
    )
    add(
        "ISSUE-HCS-861-MAJOR-001",
        "【HCS 8.6.1】【UI】健康度趋势图在空数据场景下 tooltip 残留",
        "当版本下暂无问题单数据时，趋势图空态显示正常，但鼠标悬停后仍保留上一次 tooltip 文案。该问题不绑定具体小迭代，适合测试前端对仅大版本问题单的聚合逻辑。影响范围较小，但会误导管理视图判断。",
        "HCS",
        "8.6.1",
        "",
        "关闭",
        "关闭",
        "提示",
        "UI",
        "",
        "刘畅",
        "冯璐",
        "8.6.1",
        "8.6.1",
        "8.6.1",
        "2026-03-20",
        "2026-03-22",
    )
    add(
        "ISSUE-HCSO-750-B961-001",
        "【HCSO 7.5.0-HCSO-B961】【Common】鉴权链路偶发超时",
        "离线环境下载模型时，鉴权代理偶发在 30 秒后超时返回，重试后可以恢复。排查发现 token 校验请求在高并发时会串行等待，导致整体链路抖动。问题在 HCSO 交付场景影响较大，已安排优先修复。",
        "HCSO",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B961",
        "关闭",
        "关闭",
        "严重",
        "Common",
        "REQ-HCSO-750-B961-001",
        "陈斌",
        "韩梅",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B961",
        "7.5.0-HCSO-B961",
        "2026-03-08",
        "2026-03-11",
        {"问题关闭人": "陈斌", "上线时间": "2026-03-11", "问题关闭时间": "2026-03-11"},
    )
    add(
        "ISSUE-HCSO-750-B962-001",
        "【HCSO 7.5.0-HCSO-B962】【工具】断点续传校验文件大小比对失败",
        "下载中断后重新续传，工具端完成提示正常，但最终文件大小与服务端元数据不一致。问题定位到断点续传场景下本地临时块合并顺序错误。该问题会影响离线包交付可靠性，需要在 B962 收口前解决。",
        "HCSO",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B962",
        "修复",
        "修改中",
        "一般",
        "工具",
        "REQ-HCSO-750-B962-001",
        "宋健",
        "韩梅",
        "7.5.0-HCSO",
        "7.5.0-HCSO-B962",
        "7.5.0-HCSO-B962",
        "2026-03-15",
        "2026-03-18",
    )
    add(
        "ISSUE-HC-749-B088-001",
        "【HC 7.4.9-B088】【UI】旧版本列表分页参数越界后页面无提示",
        "旧版本回归过程中，手工将分页参数改为越界值后页面未给出任何提示。接口返回了明确的错误码，但页面仅停留在加载态。该问题保留为老版本小迭代单，用于验证前端对历史版本数据的兼容展示。",
        "HC",
        "7.4.9",
        "7.4.9-B088",
        "接纳",
        "定位中",
        "提示",
        "UI",
        "",
        "刘畅",
        "冯璐",
        "7.4.9",
        "7.4.9-B088",
        "7.4.9-B088",
        "2026-03-05",
        "2026-03-19",
        {"来源系统": "Jira", "来源系统编号": "JIRA-741", "来源系统链接": "https://jira.local/browse/JIRA-741"},
    )
    add(
        "ISSUE-HC-760-B001-001",
        "【HC 7.6.0-B001】【UI】驾驶舱筛选项过多时下拉面板遮挡图表",
        "在 7.6.0 B001 版本中，团队和标签筛选项数量增加后，下拉面板会覆盖下方图表区域。问题主要出现在 1440 宽度下，移动光标后遮挡不会自动收起。该单关联当前迭代需求，便于前端验证跨版本导入后的联动查看。",
        "HC",
        "7.6.0",
        "7.6.0-B001",
        "修复完成",
        "已转测",
        "一般",
        "UI",
        "REQ-HC-760-B001-001",
        "刘畅",
        "冯璐",
        "7.6.0",
        "7.6.0-B001",
        "7.6.0-B001",
        "2026-03-30",
        "2026-04-02",
    )
    return rows


def build_timeline() -> List[Dict[str, object]]:
    data = [
        ("HC", "7.5.3", "03.03-03.09", "2026-03-09", "B001", "需求转测40%", "release/HC_7.5.3_B001", "toolkit-7.5.3.1", "2026-03-03", "2026-03-09"),
        ("HC", "7.5.3", "03.10-03.16", "2026-03-13", "B002", "需求转测80%", "release/HC_7.5.3_B002", "toolkit-7.5.3.2", "2026-03-10", "2026-03-16"),
        ("HC", "7.5.3", "03.17-03.23", "2026-03-18", "B003", "全量需求转测", "release/HC_7.5.3_B003", "toolkit-7.5.3.3", "2026-03-17", "2026-03-23"),
        ("HC", "7.5.3", "03.24-03.30", "2026-03-30", "REVIEW", "发布评审", "release/HC_7.5.3_REVIEW", "toolkit-7.5.3.4", "2026-03-24", "2026-03-30"),
        ("HC", "7.6.0", "03.24-03.30", "2026-03-30", "B001", "版本迭代", "release/HC_7.6.0_B001", "toolkit-7.6.0.1", "2026-03-24", "2026-03-30"),
        ("HC", "7.6.0", "03.31-04.06", "2026-04-02", "B002", "需求转测40%", "release/HC_7.6.0_B002", "toolkit-7.6.0.2", "2026-03-31", "2026-04-06"),
        ("HC", "7.6.0", "04.07-04.13", "2026-04-08", "B003", "需求转测80%", "release/HC_7.6.0_B003", "toolkit-7.6.0.3", "2026-04-07", "2026-04-13"),
        ("HCS", "8.6.1", "03.10-03.16", "2026-03-12", "B071", "需求转测40%", "release/HCS_8.6.1_B071", "toolkit-8.6.1.71", "2026-03-10", "2026-03-16"),
        ("HCS", "8.6.1", "03.17-03.23", "2026-03-19", "B072", "需求转测80%", "release/HCS_8.6.1_B072", "toolkit-8.6.1.72", "2026-03-17", "2026-03-23"),
        ("HCS", "8.6.1", "03.24-03.30", "2026-03-26", "REVIEW", "发布评审", "release/HCS_8.6.1_REVIEW", "toolkit-8.6.1.90", "2026-03-24", "2026-03-30"),
        ("HCSO", "7.5.0-HCSO", "03.03-03.09", "2026-03-09", "B961", "版本迭代", "release/HCSO_7.5.0_B961", "toolkit-7.5.0-hcso.961", "2026-03-03", "2026-03-09"),
        ("HCSO", "7.5.0-HCSO", "03.10-03.16", "2026-03-16", "B962", "补丁验证", "release/HCSO_7.5.0_B962", "toolkit-7.5.0-hcso.962", "2026-03-10", "2026-03-16"),
        ("HCSO", "7.5.0-HCSO", "03.17-03.23", "2026-03-23", "REVIEW", "发布评审", "release/HCSO_7.5.0_REVIEW", "toolkit-7.5.0-hcso.review", "2026-03-17", "2026-03-23"),
    ]

    rows: List[Dict[str, object]] = []
    for version_line, version_no, week, date, iteration, note, branch, toolkit, start, end in data:
        row = timeline_defaults()
        row.update(
            {
                "版本线": version_line,
                "版本号": version_no,
                "周时间": week,
                "日期": date,
                "小迭代": iteration,
                "说明": note,
                "分支": branch,
                "toolkit": toolkit,
                "开始时间": start,
                "结束时间": end,
            }
        )
        rows.append(row)
    return rows


def write_xlsx(path: Path, sheet_name: str, headers: List[str], rows: Iterable[Dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for index, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(sheet.cell(row=r, column=index).value or "")) for r in range(2, sheet.max_row + 1)))
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = min(max(max_len + 2, 12), 36)
    workbook.save(path)


def write_readme(path: Path) -> None:
    content = """版本管理双主轴驾驶舱导入示例说明

1. 需求表.xlsx
- 需求表中的“编号”为唯一主键。
- “迭代计划”使用完整小迭代编号，例如 7.5.3-B001、8.6.1-B071、7.5.0-HCSO-B961。
- “需求场景”对应版本线 HC / HCS / HCSO，“发布计划”对应版本号。

2. 问题单表.xlsx
- “关联工作项”有一部分可直接关联到需求表“编号”，也保留了少量空值用于异常数据测试。
- “交付场景”“发现版本”“发现迭代”与时间轴表保持可对应关系。
- 额外保留了仅大版本、无发现迭代、老版本小迭代等问题单示例。

3. 时间轴表.xlsx
- “版本线 + 版本号 + 小迭代”可与需求表、问题单表中的计划字段对应。
- 日期字段均写为 YYYY-MM-DD，便于前端直接解析。
"""
    path.write_text(content, encoding="utf-8")


def verify(path: Path, expected_sheet: str, expected_headers: List[str], min_rows: int) -> None:
    workbook = load_workbook(path)
    assert workbook.sheetnames == [expected_sheet], f"{path.name} sheet 名称错误"
    sheet = workbook[expected_sheet]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, len(expected_headers) + 1)]
    assert headers == expected_headers, f"{path.name} 表头不匹配"
    assert sheet.max_row - 1 >= min_rows, f"{path.name} 数据行不足"


def main() -> None:
    requirement_rows = build_requirements()
    issue_rows = build_issues()
    timeline_rows = build_timeline()

    req_path = ROOT / "需求表.xlsx"
    issue_path = ROOT / "问题单表.xlsx"
    timeline_path = ROOT / "时间轴表.xlsx"
    readme_path = ROOT / "README.txt"

    write_xlsx(req_path, "需求表", REQUIREMENT_HEADERS, requirement_rows)
    write_xlsx(issue_path, "问题单表", ISSUE_HEADERS, issue_rows)
    write_xlsx(timeline_path, "时间轴表", TIMELINE_HEADERS, timeline_rows)
    write_readme(readme_path)

    verify(req_path, "需求表", REQUIREMENT_HEADERS, 8)
    verify(issue_path, "问题单表", ISSUE_HEADERS, 8)
    verify(timeline_path, "时间轴表", TIMELINE_HEADERS, 8)

    print(f"generated: {req_path}")
    print(f"generated: {issue_path}")
    print(f"generated: {timeline_path}")
    print(f"generated: {readme_path}")


if __name__ == "__main__":
    main()
